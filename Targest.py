import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText

import Targest2

def process_tree(ws, row, col, parent=None):
    family_tree = []
    current_node = ws.cell(row=row, column=col).value

    if current_node is None or current_node.lower() == 'separator':
        return family_tree

    family_tree.append((current_node, col - 1, parent))
    child_row = row + 1
    child_col = col + 1

    while ws.cell(row=child_row, column=child_col).value is not None:
        child_node = ws.cell(row=child_row, column=child_col).value
        if child_node.lower() != 'separator':
            child_tree = process_tree(ws, child_row, child_col, current_node)
            family_tree.extend(child_tree)
        else:
            break
        child_row += 1

    sibling_col = col
    sibling_row = row + 1
    while ws.cell(row=sibling_row, column=sibling_col).value is not None:
        sibling_node = ws.cell(row=sibling_row, column=sibling_col).value
        if sibling_node.lower() != 'separator':
            sibling_tree = process_tree(ws, sibling_row, sibling_col, parent)
            family_tree.extend(sibling_tree)
        sibling_row += 1

    return family_tree

#def display_tree(tree):
#    if not tree or tree[0][0].lower() == 'separator':
#        return ""

#    tree_dict = {}
#    for node, level, parent in tree:
#        if level in tree_dict:
#            tree_dict[level].append((node, parent))
#        else:
#            tree_dict[level] = [(node, parent)]

 #   tree_str = ""

#    def print_node(node, level, last_child=False, parent=None):
#        nonlocal tree_str
#        indent = "  " * level + ("└─ " if last_child else "├─ ")
#       tree_str += indent + node + "\n"
 #       children = tree_dict.get(level + 1, [])
#        children = [child for child, child_parent in children if child_parent == node]
#        for i, child in enumerate(children):
#            print_node(child, level + 1, i == len(children) - 1, node)

#    root_node = tree_dict[0][0][0]
#    print_node(root_node, 0)
#    return tree_str

"""
def Tree():
    # Load the workbook and select the worksheet
    workbook = openpyxl.load_workbook('family_trees.xlsx')
    ws = workbook.active

    row = 2
    family_trees = []

    while row < ws.max_row:
        tree = process_tree(ws, row, 1, None)  # Pass None as the parent for the root node
        if tree:
            family_trees.append(tree)
            row += len(tree)
        else:
            row += 1

    # Filter out trees that only have a separator
    family_trees = [tree for tree in family_trees if not (len(tree) == 1 and tree[0][0].lower() == 'separator')]

    # Return the family trees
    return family_trees
"""

def build_tree_structure(tree):
    tree_structure = {}
    for line in tree:
        if len(line) != 3:
            continue

        node, col, parent = line
        level = col - 1  # Change this line
        if level not in tree_structure:
            tree_structure[level] = []
        tree_structure[level].append({"node": node, "parent": parent})

    return tree_structure


def display_tree3(tree_structure):
    if not tree_structure or 0 not in tree_structure:
        return ""

    tree_str = ""
    root_node = tree_structure[0][0]

    def print_node(node, level, last_child=False, parent=None):
        nonlocal tree_str
        indent = "│   " * level if level > 0 else ""
        indent += "└─ " if last_child else "├─ "
        tree_str += indent + node["node"] + "\n"
        children = tree_structure.get(level + 1, [])
        children = [child for child in children if child["parent"] == node["node"]]
        for i, child in enumerate(children):
            print_node(child, level + 1, i == len(children) - 1, node["node"])

    print_node(root_node, 0)
    return tree_str




def text3(window):
    family_trees = Targest2.guiTree()
    print("text3 function called") 
    print(family_trees)
    data_string = convert_to_string(family_trees)
    print(data_string)
    scrolled_text_box = ScrolledText(window, wrap=tk.WORD, height=15, width=45)
    scrolled_text_box.place(x=610, y=240)
    scrolled_text_box.configure(bg='grey', fg='white')
    #scrolled_text_box.delete(1.0, tk.END)  # Clear the scrolltext box
    #scrolled_text_box.insert(tk.END, data_string)  # Insert the converted string
    #scrolled_text_box.insert(tk.END, f"Family Tree {i}:\n")
    #scrolled_text_box.insert(tk.END, "------------\n")
    #scrolled_text_box.insert(tk.END, display_result + "\n")
    scrolled_text_box.insert(tk.END, data_string)  # Insert the converted string
    #for i, tree in enumerate(family_trees, 1):
     #   display_result = display_tree2(tree)
      #  if display_result:
       #     scrolled_text_box.insert(tk.END, f"Family Tree {i}:\n")
        #    scrolled_text_box.insert(tk.END, "------------\n")
         #   #scrolled_text_box.insert(tk.END, display_result + "\n")
          #  scrolled_text_box.insert(tk.END, data_string)  # Insert the converted string

def text2(window):
    family_trees = Targest2.guiTree()
    scrolled_text_box = ScrolledText(window, wrap=tk.WORD, height=15, width=30)
    scrolled_text_box.place(x=610, y=240)
    scrolled_text_box.configure(bg='grey', fg='white')
    for i, tree in enumerate(family_trees, 1):
        tree_structure = build_tree_structure(tree)
        display_result = display_tree3(tree_structure)
        if display_result:
            scrolled_text_box.insert(tk.END, f"Family Tree {i}:\n")
            scrolled_text_box.insert(tk.END, "------------\n")
            scrolled_text_box.insert(tk.END, display_result + "\n")

def display_tree2(tree):
    if not tree or tree[0][0].lower() == 'separator':
        return ""

    tree_dict = {}
    for node, level, parent in tree:
        if level in tree_dict:
            tree_dict[level].append((node, parent))
        else:
            tree_dict[level] = [(node, parent)]

    tree_str = ""

    def print_node(node, level, last_child=False, parent=None):
        nonlocal tree_str
        indent = "  " * level + ("└─ " if last_child else "├─ ")
        tree_str += indent + node + "\n"
        children = tree_dict.get(level + 1, [])
        children = [child for child, child_parent in children if child_parent == node]
        for i, child in enumerate(children):
            print_node(child, level + 1, i == len(children) - 1, node)

    root_node = tree_dict[0][0][0]
    print_node(root_node, 0)
    return tree_str

def convert_to_string(data):
    result = ""
    for inner_list in data:
        for item in inner_list:
            result += item
        result += "\n"
    return result
